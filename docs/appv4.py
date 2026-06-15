"""
app.py
AI Month-End Close Reconciliation Assistant — Streamlit Frontend
Run: streamlit run app.py
"""

import streamlit as st
import pandas as pd
import plotly.graph_objects as go
from io import BytesIO
from datetime import datetime
from reconciliation_engine import run_analysis

# ═══════════════════════════════════════════════════════════
# PAGE
# ═══════════════════════════════════════════════════════════

st.set_page_config(page_title="AI Month-End Close Assistant",
                   page_icon="📊", layout="wide",
                   initial_sidebar_state="expanded")

# ═══════════════════════════════════════════════════════════
# CSS
# ═══════════════════════════════════════════════════════════

st.markdown("""
<style>
.block-container{padding-top:1.4rem;padding-bottom:2rem}
hr{margin:1.2rem 0;border-color:#eaeaea}
div[data-testid="stMetric"]{
  background:#fff;border:1px solid #e2e6ea;border-radius:10px;
  padding:14px 18px;box-shadow:0 1px 3px rgba(0,0,0,.04);
}
div[data-testid="stMetric"] label{
  font-size:.72rem!important;color:#6c757d!important;
  text-transform:uppercase;letter-spacing:.03em;
}
div[data-testid="stMetric"] [data-testid="stMetricValue"]{
  font-size:1.4rem!important;font-weight:700!important;
}
.sh{font-size:.92rem;font-weight:700;color:#1a1a2e;padding-bottom:.3rem;
  border-bottom:2px solid #1a1a2e;margin-bottom:.75rem;margin-top:.2rem;
  letter-spacing:.025em;text-transform:uppercase}
.vp{background:#fff;border:1px solid #e2e6ea;border-radius:10px;
  padding:20px 22px;box-shadow:0 1px 4px rgba(0,0,0,.04);height:100%}
.vp .vp-icon{font-size:1.6rem;margin-bottom:8px}
.vp .vp-title{font-size:.88rem;font-weight:700;color:#1a1a2e;margin-bottom:6px}
.vp .vp-desc{font-size:.78rem;color:#555;line-height:1.55}
.risk-banner{border-radius:10px;padding:20px 28px;margin-bottom:4px;
  display:flex;flex-wrap:wrap;align-items:center;gap:28px}
.risk-banner .rb-level{font-size:1.1rem;font-weight:800;letter-spacing:.04em;
  text-transform:uppercase;min-width:180px}
.risk-banner .rb-stats{display:flex;flex-wrap:wrap;gap:24px}
.risk-banner .rb-stat{text-align:center}
.risk-banner .rb-stat .rb-val{font-size:1.15rem;font-weight:700;display:block}
.risk-banner .rb-stat .rb-lbl{font-size:.68rem;text-transform:uppercase;
  letter-spacing:.03em;opacity:.75}
.rb-high{background:#fbeaea;border:1px solid #e8c4c4;color:#721c24}
.rb-med{background:#fef9ea;border:1px solid #e8ddb3;color:#856404}
.rb-low{background:#e9f7ee;border:1px solid #bce0c9;color:#155724}
.ac{background:#fff;border:1px solid #e2e6ea;border-radius:8px;
  padding:14px 16px;box-shadow:0 1px 3px rgba(0,0,0,.04);
  margin-bottom:8px;height:100%}
.ac .ac-num{font-size:.65rem;font-weight:700;color:#6c757d;
  text-transform:uppercase;letter-spacing:.04em;margin-bottom:4px}
.ac .badge{display:inline-block;font-size:.6rem;font-weight:700;
  padding:2px 7px;border-radius:3px;letter-spacing:.04em;
  text-transform:uppercase;margin-left:6px;vertical-align:middle}
.badge-high{background:#f8d7da;color:#721c24}
.badge-medium{background:#fff3cd;color:#856404}
.badge-low{background:#d4edda;color:#155724}
.ac .ac-title{font-size:.84rem;font-weight:700;color:#1a1a2e;
  margin-bottom:3px;line-height:1.25}
.ac .ac-meta{font-size:.73rem;color:#6c757d;margin-bottom:6px}
.ac .ac-action{font-size:.78rem;color:#333;line-height:1.45}
.ac .ac-je{font-size:.68rem;color:#888;margin-top:5px}
</style>
""", unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════
# PLOTLY
# ═══════════════════════════════════════════════════════════

COLORS = {
    "DUPLICATE_BILL":"#c0392b","UNLINKED_ADDITION":"#e67e22",
    "MISSING_ORIGINAL_BILL":"#8e44ad","MISSING_AMORTIZATION_JE":"#2980b9",
    "OVER_AMORTIZATION":"#d35400","UNDER_AMORTIZATION":"#27ae60",
    "MANUAL_ADJUSTMENT":"#7f8c8d",
}
PL = dict(paper_bgcolor="rgba(0,0,0,0)",plot_bgcolor="rgba(0,0,0,0)",
          font=dict(family="Inter,Segoe UI,sans-serif",size=12,color="#333"),
          margin=dict(l=0,r=0,t=40,b=0),
          legend=dict(orientation="h",y=-.18,x=.5,xanchor="center"))

# Tab colors for Excel export
TAB_COLORS = {
    "Executive Summary": "1a1a2e",
    "Action Register":   "c0392b",
    "Reconciliation":    "2980b9",
    "Exceptions":        "e67e22",
    "Investigation":     "8e44ad",
    "Journal Entries":   "27ae60",
    "Input Validation":  "7f8c8d",
}


# ═══════════════════════════════════════════════════════════
# EXCEL EXPORT BUILDER
# ═══════════════════════════════════════════════════════════

def build_excel_package(results, risk_label, focus_text):
    """Build a multi-sheet Excel close review package in memory.

    Returns BytesIO buffer ready for st.download_button.
    """
    buf = BytesIO()

    recon   = results["reconciliation_summary"]
    exc     = results["exceptions"]
    inv     = results["investigation_report"]
    jnl     = results["journal_entries"]
    info    = results["validation_info"]
    counts  = results["category_counts"]
    flagged = results["flagged_transactions"]

    n_total  = len(recon)
    n_recon  = int((recon["status"] == "RECONCILED").sum())
    n_exc    = len(exc)
    rate     = (n_recon / n_total * 100) if n_total > 0 else 0.0
    t_var    = recon["abs_variance"].sum()
    n_high   = int((inv["severity"] == "HIGH").sum()) if not inv.empty else 0
    c_impact = inv["estimated_impact"].sum() if not inv.empty else 0.0
    n_find   = len(inv)
    top_issue = ""
    if not inv.empty:
        top_row = inv.loc[inv["estimated_impact"].idxmax()]
        top_issue = f"{top_row['issue_title']} ({top_row['entity']}, ${top_row['estimated_impact']:,.2f})"

    je_ids = set(jnl["root_cause_id"].tolist()) if not jnl.empty else set()

    with pd.ExcelWriter(buf, engine="openpyxl") as writer:

        # ── Helper: style a worksheet ─────────────────────────
        def _style_sheet(ws, sheet_name, dollar_cols=None, freeze_row=2):
            from openpyxl.styles import Font, Alignment, PatternFill, numbers
            from openpyxl.utils import get_column_letter

            # Tab color
            if sheet_name in TAB_COLORS:
                ws.sheet_properties.tabColor = TAB_COLORS[sheet_name]

            # Freeze panes
            ws.freeze_panes = ws.cell(row=freeze_row, column=1)

            # Bold header row
            header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=10)
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Auto-width + dollar formatting
            for col_idx, col_cells in enumerate(ws.columns, 1):
                max_len = 0
                col_letter = get_column_letter(col_idx)
                header_val = str(ws.cell(row=1, column=col_idx).value or "")

                for cell in col_cells:
                    val = str(cell.value) if cell.value is not None else ""
                    max_len = max(max_len, len(val))

                    # Dollar format
                    if dollar_cols and header_val in dollar_cols:
                        if cell.row > 1 and isinstance(cell.value, (int, float)):
                            cell.number_format = '$#,##0.00'

                ws.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 45)

        # ── Sheet 1: Executive Summary ────────────────────────
        exec_data = [
            ["Close Review Package", ""],
            ["Generated", datetime.now().strftime("%Y-%m-%d %H:%M")],
            ["", ""],
            ["EXECUTIVE SUMMARY", ""],
            ["Close Risk Level", risk_label],
            ["Date Range", f"{info['date_range_start']} → {info['date_range_end']}"],
            ["Entities", info["entities"]],
            ["Periods Analyzed", n_total],
            ["", ""],
            ["RECONCILIATION RESULTS", ""],
            ["Reconciled Periods", n_recon],
            ["Exception Periods", n_exc],
            ["Success Rate", f"{rate:.1f}%"],
            ["Total Unreconciled Variance", t_var],
            ["", ""],
            ["ROOT CAUSE ANALYSIS", ""],
            ["Root Causes Identified", n_find],
            ["High-Risk Issues", n_high],
            ["Corrective Impact Identified", c_impact],
            ["Top Issue by Impact", top_issue],
            ["", ""],
            ["MANAGEMENT FOCUS", ""],
            ["Recommended Actions", focus_text],
            ["", ""],
            ["INPUT DATA", ""],
            ["Ledger Rows", info["ledger_rows"]],
            ["Ledger Columns", info["ledger_cols"]],
            ["Amortization Rows", info["amort_rows"]],
            ["Amortization Columns", info["amort_cols"]],
            ["", ""],
            ["", "Export generated by AI Month-End Close Reconciliation Assistant"],
        ]
        exec_df = pd.DataFrame(exec_data, columns=["Field", "Value"])
        exec_df.to_excel(writer, sheet_name="Executive Summary", index=False)

        ws = writer.sheets["Executive Summary"]
        if "Executive Summary" in TAB_COLORS:
            ws.sheet_properties.tabColor = TAB_COLORS["Executive Summary"]
        ws.freeze_panes = ws.cell(row=2, column=1)
        ws.column_dimensions["A"].width = 32
        ws.column_dimensions["B"].width = 55

        from openpyxl.styles import Font, Alignment, PatternFill, numbers

        # Style header
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # Style section headers
        section_rows = [1, 4, 10, 16, 22, 25]
        for r in section_rows:
            cell = ws.cell(row=r + 1, column=1)  # +1 because row 1 is the df header
            if cell.value:
                cell.font = Font(bold=True, size=11, color="1a1a2e")

        # Dollar format for specific value cells
        for row in ws.iter_rows(min_row=2, max_col=2):
            label = str(row[0].value or "")
            if any(kw in label for kw in ["Variance", "Impact", "Corrective"]):
                if isinstance(row[1].value, (int, float)):
                    row[1].number_format = '$#,##0.00'

        # Footer note styling
        last_row = ws.max_row
        ws.cell(row=last_row, column=2).font = Font(italic=True, size=9, color="888888")

        # ── Sheet 2: Action Register ──────────────────────────
        if not inv.empty:
            def _short(row):
                cat = row["root_cause_category"]
                item = str(row.get("evidence_prepaid_item_id", ""))
                m = {
                    "DUPLICATE_BILL": "Confirm with AP and reverse duplicate",
                    "UNLINKED_ADDITION": "Create schedule or reclassify to expense",
                    "MISSING_ORIGINAL_BILL": f"Locate original invoice for {item}",
                    "MISSING_AMORTIZATION_JE": f"Post catch-up amortization for {item}",
                    "OVER_AMORTIZATION": f"Reverse excess amortization on {item}",
                    "UNDER_AMORTIZATION": f"Post catch-up entry for {item}",
                    "MANUAL_ADJUSTMENT": "Obtain documentation and approval",
                }
                return m.get(cat, "Investigate and document")

            reg = inv[["priority_rank", "issue_title", "entity", "severity",
                       "estimated_impact", "root_cause_id"]].copy()
            reg.columns = ["Priority", "Issue", "Entity", "Severity", "Impact", "_id"]
            reg["Action Required"] = inv.apply(_short, axis=1).values
            reg["JE Available"] = reg["_id"].apply(lambda x: "Yes" if x in je_ids else "No")
            reg = reg.drop(columns="_id")
            reg.to_excel(writer, sheet_name="Action Register", index=False)
            _style_sheet(writer.sheets["Action Register"], "Action Register",
                         dollar_cols=["Impact"])
        else:
            pd.DataFrame({"Note": ["No corrective actions identified"]}).to_excel(
                writer, sheet_name="Action Register", index=False)

        # ── Sheet 3: Reconciliation Summary ───────────────────
        rcols = ["entity", "accounting_period", "expected_balance",
                 "actual_balance", "variance", "status"]
        recon[rcols].to_excel(writer, sheet_name="Reconciliation", index=False)
        _style_sheet(writer.sheets["Reconciliation"], "Reconciliation",
                     dollar_cols=["expected_balance", "actual_balance", "variance"])

        # ── Sheet 4: Exceptions ───────────────────────────────
        if not exc.empty:
            exc[rcols].to_excel(writer, sheet_name="Exceptions", index=False)
            _style_sheet(writer.sheets["Exceptions"], "Exceptions",
                         dollar_cols=["expected_balance", "actual_balance", "variance"])
        else:
            pd.DataFrame({"Note": ["No exceptions — all periods reconcile"]}).to_excel(
                writer, sheet_name="Exceptions", index=False)

        # ── Sheet 5: Investigation Report ─────────────────────
        if not inv.empty:
            inv_cols = ["priority_rank", "issue_title", "entity", "severity",
                        "estimated_impact", "impacted_periods",
                        "evidence_transactions", "why_it_matters",
                        "recommended_next_step", "executive_summary"]
            inv_out = inv[[c for c in inv_cols if c in inv.columns]].copy()
            inv_out.to_excel(writer, sheet_name="Investigation", index=False)
            _style_sheet(writer.sheets["Investigation"], "Investigation",
                         dollar_cols=["estimated_impact"])
        else:
            pd.DataFrame({"Note": ["No root causes identified"]}).to_excel(
                writer, sheet_name="Investigation", index=False)

        # ── Sheet 6: Journal Entries ──────────────────────────
        if not jnl.empty:
            jnl.to_excel(writer, sheet_name="Journal Entries", index=False)
            _style_sheet(writer.sheets["Journal Entries"], "Journal Entries",
                         dollar_cols=["estimated_impact"])
        else:
            pd.DataFrame({"Note": ["No journal entries to suggest"]}).to_excel(
                writer, sheet_name="Journal Entries", index=False)

        # ── Sheet 7: Input Validation ─────────────────────────
        val_rows = [[k, v] for k, v in info.items()]
        val_rows.append(["", ""])
        val_rows.append(["TRANSACTION CLASSIFICATION", ""])
        for cat, cnt in sorted(counts.items(), key=lambda x: -x[1]):
            val_rows.append([cat, cnt])

        val_df = pd.DataFrame(val_rows, columns=["Field", "Value"])
        val_df.to_excel(writer, sheet_name="Input Validation", index=False,
                        startrow=0)

        if not flagged.empty:
            flagged_out = flagged[["transaction_id", "accounting_period", "entity",
                                   "transaction_type", "document_number", "vendor_name",
                                   "amount", "prepaid_item_id", "transaction_category"]]
            start_row = len(val_df) + 3
            flagged_out.to_excel(writer, sheet_name="Input Validation",
                                 index=False, startrow=start_row)

        _style_sheet(writer.sheets["Input Validation"], "Input Validation")

    buf.seek(0)
    return buf


# ═══════════════════════════════════════════════════════════
# HEADER
# ═══════════════════════════════════════════════════════════

st.title("📊 AI Month-End Close Reconciliation Assistant")
st.caption("Automated prepaid expense reconciliation — root cause analysis, "
           "management action items, and corrective journal entries.")
st.divider()

# ═══════════════════════════════════════════════════════════
# SIDEBAR
# ═══════════════════════════════════════════════════════════

with st.sidebar:
    st.header("Upload Data")
    st.caption("Both files must follow the required schema.")
    ledger_file = st.file_uploader("Prepaid Expense Ledger (CSV)", type=["csv"])
    amort_file  = st.file_uploader("Amortization Schedule (CSV)", type=["csv"])
    st.divider()
    both = ledger_file is not None and amort_file is not None
    run  = st.button("▶  Run Analysis", type="primary",
                     use_container_width=True, disabled=not both)
    if not both:
        st.info("Upload both CSV files to enable analysis.", icon="📁")
    st.divider()
    st.caption("Built by Sarim Areeb")
    st.caption("Python · Pandas · Streamlit · Plotly")

# ═══════════════════════════════════════════════════════════
# EXECUTION
# ═══════════════════════════════════════════════════════════

if "results" not in st.session_state:
    st.session_state.results = None

if run and both:
    try:
        with st.spinner("Running reconciliation analysis..."):
            st.session_state.results = run_analysis(
                pd.read_csv(ledger_file), pd.read_csv(amort_file))
        st.success("Analysis complete.", icon="✅")
    except ValueError as e:
        st.error(f"Validation error: {e}", icon="🚫")
        st.session_state.results = None
    except Exception as e:
        st.error(f"Unexpected error: {e}", icon="❌")
        st.session_state.results = None

results = st.session_state.results

# ═══════════════════════════════════════════════════════════
# HERO
# ═══════════════════════════════════════════════════════════

h1, h2, h3 = st.columns(3)
h1.markdown("""
<div class="vp"><div class="vp-icon">🔍</div>
  <div class="vp-title">Detect Issues</div>
  <div class="vp-desc">Automatically identifies duplicate bills, missing
  amortization entries, unlinked prepaid additions, manual adjustments,
  and schedule mismatches.</div></div>""", unsafe_allow_html=True)
h2.markdown("""
<div class="vp"><div class="vp-icon">📐</div>
  <div class="vp-title">Quantify Impact</div>
  <div class="vp-desc">Measures unreconciled variance, estimated financial
  impact, and affected periods and entities across the close cycle.</div>
</div>""", unsafe_allow_html=True)
h3.markdown("""
<div class="vp"><div class="vp-icon">✅</div>
  <div class="vp-title">Recommend Actions</div>
  <div class="vp-desc">Generates management next steps and suggested
  corrective journal entries for every finding.</div></div>""", unsafe_allow_html=True)

st.divider()

if results is None:
    st.markdown("""
### Getting Started
1. **Upload** your GL prepaid ledger and amortization schedule using the sidebar
2. **Click** "Run Analysis" to run the full reconciliation pipeline
3. **Review** risk assessment, corrective actions, charts, and detailed findings
""")
    st.stop()

# ═══════════════════════════════════════════════════════════
# EXTRACT RESULTS
# ═══════════════════════════════════════════════════════════

recon      = results["reconciliation_summary"]
exceptions = results["exceptions"]
inv        = results["investigation_report"]
journal    = results["journal_entries"]
info       = results["validation_info"]

n_total      = len(recon)
n_reconciled = int((recon["status"] == "RECONCILED").sum())
n_exc        = len(exceptions)
rate         = (n_reconciled / n_total * 100) if n_total > 0 else 0.0
total_var    = recon["abs_variance"].sum()
n_findings   = len(inv)
n_high       = int((inv["severity"] == "HIGH").sum()) if not inv.empty else 0
corr_impact  = inv["estimated_impact"].sum() if not inv.empty else 0.0
je_ids       = set(journal["root_cause_id"].tolist()) if not journal.empty else set()

# ═══════════════════════════════════════════════════════════
# 1 · CLOSE RISK BANNER
# ═══════════════════════════════════════════════════════════

if n_high >= 3 or rate < 50:
    risk_level, risk_label, risk_cls = "HIGH", "🔴 High Close Risk", "rb-high"
elif n_high >= 1 or rate < 90:
    risk_level, risk_label, risk_cls = "MEDIUM", "🟡 Elevated Close Risk", "rb-med"
else:
    risk_level, risk_label, risk_cls = "LOW", "🟢 Low Close Risk", "rb-low"

st.markdown(f"""
<div class="risk-banner {risk_cls}">
  <div class="rb-level">{risk_label}</div>
  <div class="rb-stats">
    <div class="rb-stat"><span class="rb-val">${total_var:,.0f}</span><span class="rb-lbl">Unreconciled Variance</span></div>
    <div class="rb-stat"><span class="rb-val">{n_high}</span><span class="rb-lbl">High-Risk Findings</span></div>
    <div class="rb-stat"><span class="rb-val">${corr_impact:,.0f}</span><span class="rb-lbl">Corrective Impact</span></div>
    <div class="rb-stat"><span class="rb-val">{n_exc} / {n_total}</span><span class="rb-lbl">Exception Periods</span></div>
    <div class="rb-stat"><span class="rb-val">{rate:.0f}%</span><span class="rb-lbl">Success Rate</span></div>
  </div>
</div>
""", unsafe_allow_html=True)

st.write("")

# ── Management focus text (reused in export) ──────────────
_focus_verbs = []
_seen = set()
if not inv.empty:
    for _, _r in inv.head(5).iterrows():
        cat = _r["root_cause_category"]
        if cat in _seen:
            continue
        _seen.add(cat)
        _focus_verbs.append({
            "DUPLICATE_BILL": "Reverse duplicate invoice",
            "UNLINKED_ADDITION": "Create schedule or reclassify unlinked prepaid",
            "MISSING_ORIGINAL_BILL": "Locate missing original invoice",
            "MISSING_AMORTIZATION_JE": "Post missing amortization entries",
            "OVER_AMORTIZATION": "Reverse excess amortization",
            "UNDER_AMORTIZATION": "Post catch-up amortization",
            "MANUAL_ADJUSTMENT": "Review manual adjustment documentation",
        }.get(cat, "Investigate"))
        if len(_focus_verbs) == 3:
            break
focus_text = (", ".join(_focus_verbs[:-1]) + f", and {_focus_verbs[-1]}"
              if len(_focus_verbs) > 1
              else _focus_verbs[0] if _focus_verbs else "No actions required")


# ═══════════════════════════════════════════════════════════
# 2 · EXECUTIVE DASHBOARD + DOWNLOAD BUTTON
# ═══════════════════════════════════════════════════════════

st.markdown('<div class="sh">Executive Dashboard</div>', unsafe_allow_html=True)

c1, c2, c3, c4, c5, c6 = st.columns(6)
c1.metric("Unreconciled Variance", f"${total_var:,.0f}")
c2.metric("Corrective Impact",     f"${corr_impact:,.0f}")
c3.metric("High-Risk Issues",      n_high)
c4.metric("Root Causes",           n_findings)
c5.metric("Exception Periods",     f"{n_exc}/{n_total}")
c6.metric("Success Rate",          f"{rate:.0f}%")

st.write("")

# ── Download button ───────────────────────────────────────
excel_buf = build_excel_package(results, risk_label, focus_text)
timestamp = datetime.now().strftime("%Y%m%d_%H%M")

dl_col1, dl_col2, dl_col3 = st.columns([1, 2, 1])
with dl_col2:
    st.download_button(
        label="📥  Download Close Review Package",
        data=excel_buf,
        file_name=f"close_review_package_{timestamp}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.document",
        type="primary",
        use_container_width=True,
    )

st.divider()


# ═══════════════════════════════════════════════════════════
# 3 · TOP CORRECTIVE ACTIONS REQUIRED
# ═══════════════════════════════════════════════════════════

st.markdown('<div class="sh">Top Corrective Actions Required</div>', unsafe_allow_html=True)

if inv.empty:
    st.success("No issues requiring corrective action.", icon="✅")
else:
    def _short_action(row):
        cat = row["root_cause_category"]
        item = str(row.get("evidence_prepaid_item_id", ""))
        txns = str(row.get("evidence_transactions", ""))
        has_txn = txns not in ("nan", "", "No JE found in ledger", "None found in ledger")
        m = {
            "DUPLICATE_BILL": "Confirm with AP and reverse duplicate" + (f" ({txns})" if has_txn else ""),
            "UNLINKED_ADDITION": "Create amortization schedule or reclassify to expense",
            "MISSING_ORIGINAL_BILL": f"Locate original invoice for {item}",
            "MISSING_AMORTIZATION_JE": f"Post catch-up amortization for {item}",
            "OVER_AMORTIZATION": f"Reverse excess amortization on {item}",
            "UNDER_AMORTIZATION": f"Post catch-up entry for {item}",
            "MANUAL_ADJUSTMENT": "Obtain documentation and approval" + (f" for {txns}" if has_txn else ""),
        }
        return m.get(cat, "Investigate and document")

    top5 = inv.head(5)
    row1 = top5.iloc[:3]
    cols1 = st.columns(3)
    for idx, (_, item) in enumerate(row1.iterrows()):
        sev = item["severity"]
        icon = {"HIGH": "🔴", "MEDIUM": "🟡", "LOW": "🟢"}.get(sev, "⚪")
        bcls = f"badge-{sev.lower()}"
        action = _short_action(item)
        has_je = "Yes" if item["root_cause_id"] in je_ids else "No"
        cols1[idx].markdown(f"""
        <div class="ac">
          <div class="ac-num">Action {item['priority_rank']}<span class="badge {bcls}">{icon} {sev}</span></div>
          <div class="ac-title">{item['issue_title']}</div>
          <div class="ac-meta">{item['entity']}  ·  ${item['estimated_impact']:,.2f}</div>
          <div class="ac-action">{action}</div>
          <div class="ac-je">Suggested JE available: {has_je}</div>
        </div>""", unsafe_allow_html=True)

    if len(top5) > 3:
        row2 = top5.iloc[3:]
        cols2 = st.columns(3)
        for idx, (_, item) in enumerate(row2.iterrows()):
            sev = item["severity"]
            icon = {"HIGH": "🔴", "MEDIUM": "🟡", "LOW": "🟢"}.get(sev, "⚪")
            bcls = f"badge-{sev.lower()}"
            action = _short_action(item)
            has_je = "Yes" if item["root_cause_id"] in je_ids else "No"
            cols2[idx].markdown(f"""
            <div class="ac">
              <div class="ac-num">Action {item['priority_rank']}<span class="badge {bcls}">{icon} {sev}</span></div>
              <div class="ac-title">{item['issue_title']}</div>
              <div class="ac-meta">{item['entity']}  ·  ${item['estimated_impact']:,.2f}</div>
              <div class="ac-action">{action}</div>
              <div class="ac-je">Suggested JE available: {has_je}</div>
            </div>""", unsafe_allow_html=True)

st.divider()


# ═══════════════════════════════════════════════════════════
# 4 · ROOT CAUSE ANALYTICS
# ═══════════════════════════════════════════════════════════

st.markdown('<div class="sh">Root Cause Analytics</div>', unsafe_allow_html=True)

if inv.empty or n_findings == 0:
    st.info("No root cause data available for charts.", icon="ℹ️")
else:
    ch1, ch2 = st.columns(2)

    with ch1:
        cat_df = (inv.groupby("root_cause_category")["estimated_impact"]
                  .sum().reset_index().sort_values("estimated_impact", ascending=True))
        cat_df["color"] = cat_df["root_cause_category"].map(COLORS).fillna("#95a5a6")
        cat_df["label"] = cat_df["root_cause_category"].str.replace("_", " ").str.title()
        fig1 = go.Figure(go.Bar(
            x=cat_df["estimated_impact"], y=cat_df["label"],
            orientation="h", marker_color=cat_df["color"],
            text=cat_df["estimated_impact"].apply(lambda v: f"${v:,.0f}"),
            textposition="outside", textfont=dict(size=11)))
        fig1.update_layout(
            **PL, title=dict(text="Financial Impact by Root Cause", font=dict(size=13)),
            xaxis=dict(title="", showgrid=True, gridcolor="#f0f0f0",
                       zeroline=False, tickprefix="$", tickformat=","),
            yaxis=dict(title=""), height=320, showlegend=False)
        st.plotly_chart(fig1, use_container_width=True)

    with ch2:
        ent_df = (inv.groupby("entity")["estimated_impact"]
                  .sum().reset_index().sort_values("estimated_impact", ascending=True))
        fig2 = go.Figure(go.Bar(
            x=ent_df["estimated_impact"], y=ent_df["entity"],
            orientation="h", marker_color="#2980b9",
            text=ent_df["estimated_impact"].apply(lambda v: f"${v:,.0f}"),
            textposition="outside", textfont=dict(size=11)))
        fig2.update_layout(
            **PL, title=dict(text="Financial Impact by Entity", font=dict(size=13)),
            xaxis=dict(title="", showgrid=True, gridcolor="#f0f0f0",
                       zeroline=False, tickprefix="$", tickformat=","),
            yaxis=dict(title=""), height=320, showlegend=False)
        st.plotly_chart(fig2, use_container_width=True)

st.divider()


# ═══════════════════════════════════════════════════════════
# 5 · CORRECTIVE ACTION REGISTER
# ═══════════════════════════════════════════════════════════

st.markdown('<div class="sh">Corrective Action Register</div>', unsafe_allow_html=True)

if inv.empty:
    st.info("No corrective actions to display.", icon="ℹ️")
else:
    register = inv[["priority_rank", "issue_title", "entity", "severity",
                     "estimated_impact", "root_cause_id"]].copy()
    register.columns = ["Priority", "Issue", "Entity", "Severity", "Impact", "_rc_id"]
    register["Action Required"] = inv.apply(_short_action, axis=1).values
    register["JE Available"] = register["_rc_id"].apply(
        lambda x: "✅ Yes" if x in je_ids else "—")
    register = register.drop(columns="_rc_id")

    def _hl_sev_reg(val):
        if val == "HIGH":   return "background-color:#f8d7da;font-weight:bold;color:#721c24"
        if val == "MEDIUM": return "background-color:#fff3cd;font-weight:bold;color:#856404"
        if val == "LOW":    return "background-color:#d4edda;color:#155724"
        return ""

    st.dataframe(
        register.style.format({"Impact": "${:,.2f}"})
        .map(_hl_sev_reg, subset=["Severity"]),
        use_container_width=True, hide_index=True)

st.divider()


# ═══════════════════════════════════════════════════════════
# 6 · DETAILED ANALYSIS (tabs)
# ═══════════════════════════════════════════════════════════

st.markdown('<div class="sh">Detailed Analysis</div>', unsafe_allow_html=True)

tab_val, tab_rec, tab_exc, tab_inv, tab_je = st.tabs([
    "📁 Input Validation", "📊 Reconciliation", "⚠️ Exceptions",
    "🔍 Investigation Report", "📝 Journal Entries"])

with tab_val:
    v1, v2, v3, v4 = st.columns(4)
    v1.metric("Ledger Rows", f"{info['ledger_rows']:,}")
    v2.metric("Schedule Rows", f"{info['amort_rows']:,}")
    v3.metric("Entities", info["entities"])
    v4.metric("Periods", info["periods"])
    st.write(f"**Date range:** {info['date_range_start']} → {info['date_range_end']}  ·  "
             f"**Ledger cols:** {info['ledger_cols']}  ·  **Schedule cols:** {info['amort_cols']}")
    st.write("**Transaction classification:**")
    counts = results["category_counts"]
    st.dataframe(pd.DataFrame([{"Category": k, "Count": v}
        for k, v in sorted(counts.items(), key=lambda x: -x[1])]),
        use_container_width=True, hide_index=True)
    flagged = results["flagged_transactions"]
    if not flagged.empty:
        st.write(f"**Flagged for review:** {len(flagged)} row(s)")
        st.dataframe(
            flagged[["transaction_id", "accounting_period", "entity",
                     "transaction_type", "document_number", "vendor_name",
                     "amount", "prepaid_item_id", "transaction_category"]]
            .style.format({"amount": "${:,.2f}"}),
            use_container_width=True, hide_index=True)

with tab_rec:
    def _hl_st(v):
        if v == "RECONCILED": return "background-color:#d4edda;color:#155724;font-weight:bold"
        if v == "EXCEPTION":  return "background-color:#f8d7da;color:#721c24;font-weight:bold"
        return ""
    rcols = ["entity", "accounting_period", "expected_balance",
             "actual_balance", "variance", "status"]
    st.dataframe(recon[rcols].style
        .format({"expected_balance": "${:,.2f}", "actual_balance": "${:,.2f}", "variance": "${:,.2f}"})
        .map(_hl_st, subset=["status"]),
        use_container_width=True, hide_index=True)

with tab_exc:
    if exceptions.empty:
        st.success("All periods reconcile cleanly.", icon="✅")
    else:
        st.warning(f"{n_exc} exception period(s) · ${total_var:,.2f} total variance", icon="⚠️")
        st.dataframe(exceptions[rcols].style
            .format({"expected_balance": "${:,.2f}", "actual_balance": "${:,.2f}", "variance": "${:,.2f}"})
            .map(_hl_st, subset=["status"]),
            use_container_width=True, hide_index=True)

with tab_inv:
    if inv.empty:
        st.info("No root causes identified.", icon="ℹ️")
    else:
        def _hl_sv(v):
            if v == "HIGH":   return "background-color:#f8d7da;font-weight:bold;color:#721c24"
            if v == "MEDIUM": return "background-color:#fff3cd;font-weight:bold;color:#856404"
            if v == "LOW":    return "background-color:#d4edda;color:#155724"
            return ""
        st.dataframe(
            inv[["priority_rank", "issue_title", "entity", "severity",
                 "estimated_impact", "recommended_next_step"]].style
            .format({"estimated_impact": "${:,.2f}"})
            .map(_hl_sv, subset=["severity"]),
            use_container_width=True, hide_index=True)
        st.markdown("#### Detailed Findings")
        for _, row in inv.iterrows():
            si = {"HIGH": "🔴", "MEDIUM": "🟡", "LOW": "🟢"}.get(row["severity"], "⚪")
            with st.expander(f"{si}  #{row['priority_rank']}  {row['issue_title']}  —  "
                             f"{row['entity']}  ·  ${row['estimated_impact']:,.2f}"):
                st.markdown(f"**Severity:** {row['severity']}  ·  "
                            f"**Impact:** ${row['estimated_impact']:,.2f}  ·  "
                            f"**Entity:** {row['entity']}")
                st.markdown(f"**Periods affected:** {row['impacted_periods']}")
                st.markdown(f"**Evidence:** {row['evidence_transactions']}")
                st.markdown("---")
                st.markdown("**Why this matters**")
                st.markdown(row["why_it_matters"])
                st.markdown("**Recommended next step**")
                st.markdown(row["recommended_next_step"])
                st.markdown("**Executive summary**")
                st.markdown(row["executive_summary"])

with tab_je:
    if journal.empty:
        st.info("No corrective journal entries to suggest.", icon="ℹ️")
    else:
        st.write(f"{len(journal)} suggested entries · "
                 f"${journal['estimated_impact'].sum():,.2f} total impact")
        def _hl_je(v):
            if v == "HIGH":   return "background-color:#f8d7da;font-weight:bold;color:#721c24"
            if v == "MEDIUM": return "background-color:#fff3cd;font-weight:bold;color:#856404"
            if v == "LOW":    return "background-color:#d4edda;color:#155724"
            return ""
        st.dataframe(
            journal[["root_cause_id", "entity", "category", "severity",
                      "estimated_impact", "evidence_prepaid_item_id"]].style
            .format({"estimated_impact": "${:,.2f}"})
            .map(_hl_je, subset=["severity"]),
            use_container_width=True, hide_index=True)
        st.markdown("#### Entry Details")
        for _, row in journal.iterrows():
            with st.expander(f"{row['root_cause_id']}  ·  {row['category']}  ·  "
                             f"{row['entity']}  ·  ${row['estimated_impact']:,.2f}"):
                st.markdown(f"**Item:** {row['evidence_prepaid_item_id']}")
                st.markdown("**Suggested journal entry:**")
                st.code(row["suggested_journal_entry"], language=None)
                st.markdown("**Recommended action:**")
                st.markdown(row["recommended_action"])

# ═══════════════════════════════════════════════════════════
# FOOTER
# ═══════════════════════════════════════════════════════════

st.divider()
st.caption("AI Month-End Close Reconciliation Assistant  ·  "
           "Built by Sarim Areeb  ·  "
           "Python · Pandas · Streamlit · Plotly  ·  "
           "Backend: reconciliation_engine.py")
